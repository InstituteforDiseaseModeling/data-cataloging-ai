"""
generate_databio.py

Reads databio_draft.json from the current working directory and produces
a formatted DataBio Excel file.

Usage:
    python generate_databio.py
    python generate_databio.py --input path/to/databio_draft.json
    python generate_databio.py --input draft.json --output MyDataset_DataBio.xlsx
"""

import json
import sys
import os
from datetime import date
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CLI args ───────────────────────────────────────────────────────────────────
args = sys.argv[1:]
input_path  = "databio_draft.json"
output_path = None

i = 0
while i < len(args):
    if args[i] == "--input"  and i + 1 < len(args): input_path  = args[i+1]; i += 2
    elif args[i] == "--output" and i + 1 < len(args): output_path = args[i+1]; i += 2
    else: i += 1

if not Path(input_path).exists():
    sys.exit(f"ERROR: {input_path} not found. Run the complete-databio skill first to generate it.")

with open(input_path, encoding="utf-8") as f:
    data = json.load(f)

dataset_name   = data.get("dataset_name", "Dataset")
output_path    = output_path or data.get("output_filename") or f"{dataset_name.replace(' ','_')}_DataBio.xlsx"
generated_date = data.get("generated_date", str(date.today()))
mode           = data.get("mode", "")
sources_used   = data.get("sources_used", [])
metadata_rows  = data.get("metadata", [])
databio_rows   = data.get("data_bio", [])
variable_rows  = data.get("variables", [])

# ── Shared styles ──────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
SUBHDR_FILL  = PatternFill("solid", fgColor="2E75B6")
ALT_FILL     = PatternFill("solid", fgColor="D6E4F0")
REVIEW_FILL  = PatternFill("solid", fgColor="FFE699")
SENSITIVE_FILL = PatternFill("solid", fgColor="FFB3B3")
WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")

HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
SUBHDR_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
BOLD_FONT    = Font(bold=True, name="Calibri", size=10)
BOLD_BLUE    = Font(bold=True, name="Calibri", size=10, color="1F4E79")
NORMAL_FONT  = Font(name="Calibri", size=10)

THIN  = Side(style="thin",   color="BFBFBF")
THICK = Side(style="medium", color="2E75B6")
def thin_border():  return Border(left=THIN,  right=THIN,  top=THIN,  bottom=THIN)
def thick_border(): return Border(left=THICK, right=THICK, top=THICK, bottom=THICK)

WRAP        = Alignment(wrap_text=True, vertical="top")
CTR_WRAP    = Alignment(wrap_text=True, vertical="top", horizontal="center")

def set_header_row(ws, row, ncols, fill=HEADER_FILL, font=HEADER_FONT):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill, cell.font, cell.alignment, cell.border = fill, font, CTR_WRAP, thin_border()

def set_data_row(ws, row, ncols, alt=False, review_cols=(), sensitive_cols=()):
    base = ALT_FILL if alt else WHITE_FILL
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        val  = cell.value
        if c in sensitive_cols and val is True:
            cell.fill = SENSITIVE_FILL
            cell.font = Font(bold=True, name="Calibri", size=10)
        elif c in review_cols and val is True:
            cell.fill = REVIEW_FILL
            cell.font = NORMAL_FONT
        else:
            cell.fill = base
            cell.font = NORMAL_FONT
        cell.alignment = WRAP
        cell.border     = thin_border()

def set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

def banner(ws, row, ncols, text, font_size=13):
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=font_size)
    c.fill      = HEADER_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 30

wb = Workbook()

# ══════════════════════════════════════════════════════════════════════════════
# LEGEND sheet (index 0)
# ══════════════════════════════════════════════════════════════════════════════
ws_legend = wb.active
ws_legend.title = "LEGEND"
banner(ws_legend, 1, 3, f"DataBio — {dataset_name}   |   Generated: {generated_date}")

legend_items = [
    ("", "Field", "Description"),
    (REVIEW_FILL,    "Yellow cell",  "Needs human review — field was not auto-filled or requires confirmation"),
    (SENSITIVE_FILL, "Red cell",     "Sensitive variable — contains or may contain personal, health, or restricted data"),
    (ALT_FILL,       "Light blue row", "Alternating row shading for readability"),
    (SUBHDR_FILL,    "Dark blue row", "Section header — DATA BIO section divider"),
]
ws_legend.cell(row=3, column=2, value="Color / style").font = BOLD_FONT
ws_legend.cell(row=3, column=3, value="Meaning").font = BOLD_FONT
for r, (fill, label, desc) in enumerate(legend_items[1:], 4):
    c1 = ws_legend.cell(row=r, column=2, value=label)
    c1.fill, c1.font, c1.alignment, c1.border = fill, NORMAL_FONT, CTR_WRAP, thin_border()
    c2 = ws_legend.cell(row=r, column=3, value=desc)
    c2.font, c2.alignment, c2.border = NORMAL_FONT, WRAP, thin_border()

# Sources used
if sources_used:
    ws_legend.cell(row=10, column=2, value="Sources used to generate this draft:").font = BOLD_FONT
    for j, src in enumerate(sources_used, 11):
        ws_legend.cell(row=j, column=2, value=f"• {src}").font = NORMAL_FONT
    note_row = 11 + len(sources_used)
else:
    note_row = 11

if mode:
    ws_legend.cell(row=note_row, column=2, value=f"Data access mode: Mode {mode}").font = BOLD_FONT

ws_legend.column_dimensions["A"].width = 3
ws_legend.column_dimensions["B"].width = 28
ws_legend.column_dimensions["C"].width = 65

# ══════════════════════════════════════════════════════════════════════════════
# METADATA sheet
# ══════════════════════════════════════════════════════════════════════════════
ws_meta = wb.create_sheet("METADATA")
banner(ws_meta, 1, 6, f"METADATA TAB — {dataset_name}")

meta_headers = ["Field", "Draft value", "Source", "Confidence", "Needs review", "Notes / human input prompt"]
for c, h in enumerate(meta_headers, 1):
    ws_meta.cell(row=2, column=c, value=h)
set_header_row(ws_meta, 2, len(meta_headers))

for i, row in enumerate(metadata_rows):
    r   = i + 3
    alt = (i % 2 == 1)
    ws_meta.cell(row=r, column=1, value=row.get("field", ""))
    ws_meta.cell(row=r, column=2, value=row.get("value", ""))
    ws_meta.cell(row=r, column=3, value=row.get("source", ""))
    ws_meta.cell(row=r, column=4, value=row.get("confidence", ""))
    ws_meta.cell(row=r, column=5, value=row.get("needs_review", False))
    ws_meta.cell(row=r, column=6, value=row.get("review_notes", ""))
    set_data_row(ws_meta, r, len(meta_headers), alt=alt, review_cols=(5,))
    ws_meta.cell(row=r, column=1).font = BOLD_FONT
    val_len = len(str(row.get("value", "")))
    ws_meta.row_dimensions[r].height = max(40, min(val_len // 2, 180))

set_col_widths(ws_meta, [28, 55, 35, 12, 14, 45])
ws_meta.freeze_panes = "B3"
ws_meta.auto_filter.ref = f"A2:F{len(metadata_rows)+2}"

# ══════════════════════════════════════════════════════════════════════════════
# DATA BIO sheet
# ══════════════════════════════════════════════════════════════════════════════
ws_bio = wb.create_sheet("DATA BIO")
banner(ws_bio, 1, 7, f"DATA BIO TAB — {dataset_name}")

bio_headers = ["Section", "Q#", "Question", "Draft response", "Source", "Confidence", "Needs review"]
for c, h in enumerate(bio_headers, 1):
    ws_bio.cell(row=2, column=c, value=h)
set_header_row(ws_bio, 2, len(bio_headers))

row_num    = 3
prev_sec   = None
for i, row in enumerate(databio_rows):
    section = row.get("section", "")
    if section != prev_sec:
        ws_bio.merge_cells(f"A{row_num}:G{row_num}")
        sc = ws_bio.cell(row=row_num, column=1, value=section)
        sc.fill, sc.font = SUBHDR_FILL, SUBHDR_FONT
        sc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        sc.border = thick_border()
        ws_bio.row_dimensions[row_num].height = 18
        row_num   += 1
        prev_sec   = section

    alt = (i % 2 == 0)
    ws_bio.cell(row=row_num, column=1, value=section)
    ws_bio.cell(row=row_num, column=2, value=row.get("question_num", ""))
    ws_bio.cell(row=row_num, column=3, value=row.get("question", ""))
    ws_bio.cell(row=row_num, column=4, value=row.get("response", ""))
    ws_bio.cell(row=row_num, column=5, value=row.get("source", ""))
    ws_bio.cell(row=row_num, column=6, value=row.get("confidence", ""))
    ws_bio.cell(row=row_num, column=7, value=row.get("needs_review", False))
    set_data_row(ws_bio, row_num, len(bio_headers), alt=alt, review_cols=(7,))
    ws_bio.cell(row=row_num, column=2).alignment = CTR_WRAP
    ws_bio.cell(row=row_num, column=6).alignment = CTR_WRAP
    ws_bio.cell(row=row_num, column=7).alignment = CTR_WRAP
    resp_len = len(str(row.get("response", "")))
    ws_bio.row_dimensions[row_num].height = max(60, min(resp_len // 2, 200))
    row_num += 1

set_col_widths(ws_bio, [28, 5, 42, 65, 30, 14, 13])
ws_bio.freeze_panes = "D3"
ws_bio.auto_filter.ref = f"A2:G{row_num-1}"

# ══════════════════════════════════════════════════════════════════════════════
# INDIVIDUAL VARIABLES sheet
# ══════════════════════════════════════════════════════════════════════════════
ws_vars = wb.create_sheet("INDIVIDUAL VARIABLES")
banner(ws_vars, 1, 13, f"INDIVIDUAL VARIABLES TAB — {dataset_name}")

var_headers = [
    "variable_name", "variable_label", "definition", "data_type", "unit",
    "allowed_values_codes", "missing_unknown_codes", "source_derivation",
    "numerator", "denominator", "sensitive", "data_quality_notes", "needs_review",
]
for c, h in enumerate(var_headers, 1):
    ws_vars.cell(row=2, column=c, value=h)
set_header_row(ws_vars, 2, len(var_headers))

SENSITIVE_COL  = 11   # column index for 'sensitive'
REVIEW_COL_VAR = 13   # column index for 'needs_review'

for i, row in enumerate(variable_rows):
    r   = i + 3
    alt = (i % 2 == 1)
    values = [
        row.get("variable_name",       ""),
        row.get("variable_label",      ""),
        row.get("definition",          ""),
        row.get("data_type",           ""),
        row.get("unit",                ""),
        row.get("allowed_values_codes",""),
        row.get("missing_unknown_codes",""),
        row.get("source_derivation",   ""),
        row.get("numerator",           ""),
        row.get("denominator",         ""),
        row.get("sensitive",           False),
        row.get("data_quality_notes",  ""),
        row.get("needs_review",        False),
    ]
    for c, v in enumerate(values, 1):
        ws_vars.cell(row=r, column=c, value=v)
    set_data_row(ws_vars, r, len(var_headers), alt=alt,
                 review_cols=(REVIEW_COL_VAR,), sensitive_cols=(SENSITIVE_COL,))
    ws_vars.cell(row=r, column=1).font = BOLD_BLUE
    def_len = len(str(row.get("definition", "")))
    ws_vars.row_dimensions[r].height = max(45, min(def_len // 2, 150))

set_col_widths(ws_vars, [22, 28, 55, 11, 18, 38, 20, 36, 32, 26, 10, 45, 13])
ws_vars.freeze_panes = "C3"
ws_vars.auto_filter.ref = f"A2:M{len(variable_rows)+2}"

# ── Save ───────────────────────────────────────────────────────────────────────
wb.save(output_path)
print(f"DataBio saved: {Path(output_path).resolve()}")
