"""
generate_catalog.py

Reads catalog_draft.json from the current working directory and fills the
catalog template files found alongside this script:

    DataProfile.xlsx  (sheets: "Metadata" and "DataBio")
    DataDict.xlsx

producing dataset-specific outputs:

    <Dataset>_DataProfile.xlsx
    <Dataset>_DataDict.xlsx

The Metadata and DataBio skills are separate, but they write into the two
sheets of the *same* DataProfile.xlsx output file. If that output file
already exists (e.g. one skill already ran), it's loaded and updated in
place rather than overwritten from the blank template, so filling one sheet
never clobbers the other.

Usage:
    python generate_catalog.py
    python generate_catalog.py --input path/to/catalog_draft.json
    python generate_catalog.py --input draft.json --output-dir out/
    python generate_catalog.py --only databio            # just the DataBio sheet
    python generate_catalog.py --only metadata,datadict   # any subset
"""

import json
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

SCRIPT_DIR = Path(__file__).resolve().parent
TEMPLATE_DATAPROFILE = SCRIPT_DIR / "DataProfile.xlsx"
TEMPLATE_DATADICT = SCRIPT_DIR / "DataDict.xlsx"

REVIEW_FILL = PatternFill("solid", fgColor="FFE699")
SENSITIVE_FILL = PatternFill("solid", fgColor="FFB3B3")
ALT_FILL = PatternFill("solid", fgColor="F2F7FC")
THIN = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")

DATADICT_KEYS = [
    "file_name", "variable_name", "variable_label", "definition", "data_type",
    "unit", "allowed_values_codes", "missing_unknown_codes", "source_derivation",
    "numerator", "denominator", "sensitive", "data_quality_notes",
]

# Metadata sheet: fields end at row 19 ("Related dataset location(s)"); row 20
# is the "To Be Completed by Modeling Technology Team" banner, and rows 21-23
# (Storage/repository location, Data steward, Data Catalog location) are that
# team's responsibility, not this skill's -- never write to those rows.
METADATA_FIRST_ROW = 3
METADATA_LAST_ROW = 19


def _review_comment(entry, fallback="Needs human review before finalizing."):
    text = entry.get("review_notes") or fallback
    return Comment(text, "Catalog generator")


def _load_or_copy_dataprofile(output_path):
    if not TEMPLATE_DATAPROFILE.exists():
        sys.exit(f"ERROR: template not found: {TEMPLATE_DATAPROFILE}")
    if output_path.exists():
        return load_workbook(output_path)
    return load_workbook(TEMPLATE_DATAPROFILE)


def fill_metadata_sheet(wb, data):
    ws = wb["Metadata"]
    entries = data.get("metadata", [])
    max_entries = METADATA_LAST_ROW - METADATA_FIRST_ROW + 1
    for i, entry in enumerate(entries[:max_entries]):
        row = METADATA_FIRST_ROW + i
        cell = ws.cell(row=row, column=2, value=entry.get("value", ""))
        if entry.get("needs_review"):
            cell.fill = REVIEW_FILL
            cell.comment = _review_comment(entry)
    return wb


def fill_databio_sheet(wb, data):
    ws = wb["DataBio"]
    for i, entry in enumerate(data.get("data_bio", [])):
        row = 3 + i
        answer_cell = ws.cell(row=row, column=4, value=entry.get("response", ""))

        notes = entry.get("source", "")
        if entry.get("needs_review"):
            flag = entry.get("review_notes") or "Needs human review."
            notes = f"⚠ Needs review — {flag}\n{notes}".strip()
            answer_cell.fill = REVIEW_FILL
            answer_cell.comment = _review_comment(entry)
        ws.cell(row=row, column=5, value=notes)
    return wb


def fill_dataprofile(data, output_dir, base_name, sheets):
    """sheets is a subset of {"metadata", "databio"} -- which sheet(s) to
    (re)fill in this run. The other sheet, if already present in an existing
    output file, is left untouched."""
    output_path = output_dir / f"{base_name}_DataProfile.xlsx"
    wb = _load_or_copy_dataprofile(output_path)

    if "metadata" in sheets:
        fill_metadata_sheet(wb, data)
    if "databio" in sheets:
        fill_databio_sheet(wb, data)

    wb.save(output_path)
    return output_path


# ── DataDict.xlsx ────────────────────────────────────────────────────────────
def fill_datadict(data, output_dir, base_name):
    if not TEMPLATE_DATADICT.exists():
        sys.exit(f"ERROR: template not found: {TEMPLATE_DATADICT}")

    wb = load_workbook(TEMPLATE_DATADICT)
    ws = wb["Sheet1"]

    variables = data.get("variables", [])
    for i, entry in enumerate(variables):
        row = 2 + i
        alt = (i % 2 == 1)
        for col, key in enumerate(DATADICT_KEYS, 1):
            cell = ws.cell(row=row, column=col, value=entry.get(key, ""))
            cell.border = THIN_BORDER
            cell.alignment = WRAP_TOP
            if alt:
                cell.fill = ALT_FILL

        if entry.get("sensitive"):
            sensitive_cell = ws.cell(row=row, column=12)
            sensitive_cell.fill = SENSITIVE_FILL
            sensitive_cell.font = Font(bold=True)

        if entry.get("needs_review"):
            name_cell = ws.cell(row=row, column=2)
            name_cell.fill = REVIEW_FILL
            name_cell.comment = _review_comment(entry)

    ws.freeze_panes = "A2"
    if variables:
        ws.auto_filter.ref = f"A1:M{len(variables) + 1}"

    output_path = output_dir / f"{base_name}_DataDict.xlsx"
    wb.save(output_path)
    return output_path


ALL_TARGETS = ("metadata", "databio", "datadict")


def main():
    args = sys.argv[1:]
    input_path = "catalog_draft.json"
    output_dir = Path.cwd()
    targets = list(ALL_TARGETS)

    i = 0
    while i < len(args):
        if args[i] == "--input" and i + 1 < len(args):
            input_path = args[i + 1]
            i += 2
        elif args[i] == "--output-dir" and i + 1 < len(args):
            output_dir = Path(args[i + 1])
            i += 2
        elif args[i] == "--only" and i + 1 < len(args):
            targets = [t.strip().lower() for t in args[i + 1].split(",")]
            unknown = [t for t in targets if t not in ALL_TARGETS]
            if unknown:
                sys.exit(f"ERROR: unknown --only target(s) {unknown}; choose from {ALL_TARGETS}")
            i += 2
        else:
            i += 1

    if not Path(input_path).exists():
        sys.exit(f"ERROR: {input_path} not found. Run the catalog-dataset skill first to generate it.")

    with open(input_path, encoding="utf-8") as f:
        data = json.load(f)

    output_dir.mkdir(parents=True, exist_ok=True)
    base_name = data.get("dataset_name", "Dataset").replace(" ", "_")

    dataprofile_sheets = {t for t in targets if t in ("metadata", "databio")}
    if dataprofile_sheets:
        path = fill_dataprofile(data, output_dir, base_name, dataprofile_sheets)
        print(f"DataProfile saved ({', '.join(sorted(dataprofile_sheets))}): {path.resolve()}")

    if "datadict" in targets:
        path = fill_datadict(data, output_dir, base_name)
        print(f"DataDict saved: {path.resolve()}")


if __name__ == "__main__":
    main()
